# -*- coding: utf-8 -*-
"""
Reports Module - Geração de relatórios em Excel
"""

import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ReportGenerator:
    """Gerador de relatórios em Excel para dimensionamento hidráulico"""
    
    @staticmethod
    def create_excel_report(filename, project_data, calculations, validations):
        """
        Cria relatório completo em Excel
        
        Args:
            filename: Nome do arquivo de saída
            project_data: Dados do projeto (vazão, trechos, etc)
            calculations: Resultados dos cálculos
            validations: Resultados das validações
        """
        wb = Workbook()
        wb.remove(wb.active)
        
        # Criar abas
        ReportGenerator._create_summary_sheet(wb, project_data, calculations, validations)
        ReportGenerator._create_segments_sheet(wb, project_data, calculations)
        ReportGenerator._create_localized_losses_sheet(wb, project_data)
        ReportGenerator._create_validation_sheet(wb, validations)
        
        wb.save(filename)
        return filename
    
    @staticmethod
    def _create_summary_sheet(wb, project_data, calculations, validations):
        """Cria aba de resumo"""
        ws = wb.create_sheet('Resumo')
        
        # Estilos
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        title_font = Font(bold=True, size=14)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws['A1'] = 'RELATÓRIO DE DIMENSIONAMENTO HIDRÁULICO'
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # Data
        ws['A2'] = f"Data: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Seção de dados de entrada
        row = 4
        ws[f'A{row}'] = 'DADOS DE ENTRADA'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        data_entries = [
            ('Vazão (m³/h)', project_data.get('flow_rate', 0)),
            ('Número de Trechos', project_data.get('num_segments', 0)),
            ('Cota Montante (m)', project_data.get('piezometric_upstream', 0)),
            ('Cota Jusante (m)', project_data.get('piezometric_downstream', 0))
        ]
        
        for label, value in data_entries:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            row += 1
        
        # Seção de resultados
        row += 1
        ws[f'A{row}'] = 'RESULTADOS PRINCIPAIS'
        ws[f'A{row}'].font = Font(bold=True, size=11)
        
        row += 1
        results = [
            ('Perda de Carga Total (m)', calculations.get('total_head_loss', 0)),
            ('Velocidade Média (m/s)', calculations.get('avg_velocity', 0)),
            ('Perda Localizada (m)', calculations.get('localized_loss', 0)),
            ('Margem Piezométrica (m)', validations.get('margin', 0))
        ]
        
        for label, value in results:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = round(value, 3) if isinstance(value, float) else value
            row += 1
        
        # Ajustar colunas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 15
    
    @staticmethod
    def _create_segments_sheet(wb, project_data, calculations):
        """Cria aba com detalhes dos trechos"""
        ws = wb.create_sheet('Trechos')
        
        header_fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # Headers
        headers = ['Trecho', 'Material', 'Diâmetro (mm)', 'Comprimento (m)', 
                   'Velocidade (m/s)', 'Perda de Carga (m)', 'Reynolds', 'Fator f']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados dos trechos
        segments = project_data.get('segments', [])
        calc_segments = calculations.get('segments', [])
        
        for row, segment in enumerate(segments, 2):
            ws.cell(row=row, column=1).value = segment.get('name', f'Trecho {row-1}') if isinstance(segment, dict) else str(segment)
            ws.cell(row=row, column=2).value = segment.get('material', '') if isinstance(segment, dict) else ''
            ws.cell(row=row, column=3).value = segment.get('diameter', 0) if isinstance(segment, dict) else 0
            ws.cell(row=row, column=4).value = segment.get('length', 0) if isinstance(segment, dict) else 0
            
            # Dados de cálculo
            if row-2 < len(calc_segments) and isinstance(calc_segments[row-2], dict):
                calc = calc_segments[row-2]
            else:
                calc = {}
            
            ws.cell(row=row, column=5).value = round(calc.get('velocity_ms', 0), 3) if isinstance(calc, dict) else 0
            ws.cell(row=row, column=6).value = round(calc.get('head_loss_m', 0), 3) if isinstance(calc, dict) else 0
            ws.cell(row=row, column=7).value = round(calc.get('reynolds', 0), 0) if isinstance(calc, dict) else 0
            ws.cell(row=row, column=8).value = round(calc.get('friction_factor', 0), 4) if isinstance(calc, dict) else 0
        
        # Ajustar colunas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _create_localized_losses_sheet(wb, project_data):
        """Cria aba com perdas localizadas"""
        ws = wb.create_sheet('Perdas Localizadas')
        
        header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        headers = ['Componente', 'Quantidade', 'K Unitário', 'K Total', 'Perda (m)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
        
        # Dados de perdas localizadas
        localized = project_data.get('localized_losses', {})
        for row, (component, data) in enumerate(localized.items(), 2):
            ws.cell(row=row, column=1).value = component
            ws.cell(row=row, column=2).value = data.get('quantity', 0)
            ws.cell(row=row, column=3).value = round(data.get('K_unit', 0), 3)
            ws.cell(row=row, column=4).value = round(data.get('K_total', 0), 3)
            ws.cell(row=row, column=5).value = round(data.get('head_loss', 0), 3)
        
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    @staticmethod
    def _create_validation_sheet(wb, validations):
        """Cria aba com validações"""
        ws = wb.create_sheet('Validações')
        
        ws['A1'] = 'VERIFICAÇÃO CONTRA NORMAS E BOAS PRÁTICAS'
        ws['A1'].font = Font(bold=True, size=12)
        
        row = 3
        
        # Validação de velocidade
        vel_val = validations.get('velocity', {})
        ws[f'A{row}'] = 'Velocidade'
        ws[f'B{row}'] = vel_val.get('status', '')
        ws[f'C{row}'] = f"{vel_val.get('velocity_ms', 0):.3f} m/s"
        row += 2
        
        # Validação de perda de carga
        hl_val = validations.get('head_loss', {})
        ws[f'A{row}'] = 'Perda de Carga'
        ws[f'B{row}'] = hl_val.get('status', '')
        ws[f'C{row}'] = f"{hl_val.get('J_m_per_100m', 0):.2f} m/100m"
        row += 2
        
        # Validação de linha piezométrica
        pz_val = validations.get('piezometric', {})
        ws[f'A{row}'] = 'Linha Piezométrica'
        ws[f'B{row}'] = pz_val.get('status', '')
        ws[f'C{row}'] = f"Margem: {pz_val.get('margin', 0):.3f}m"
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 25
